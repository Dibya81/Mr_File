from openpyxl import load_workbook


class ExcelExtractor:
    def extract(self, file_path: str) -> dict:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)

            sheet_count = len(wb.sheetnames)
            sheet_names = wb.sheetnames

            total_rows = 0
            total_cols = 0
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                if ws.max_row:
                    total_rows += ws.max_row
                if ws.max_column:
                    total_cols = max(total_cols, ws.max_column)

            result = {
                "sheet_count": sheet_count,
                "sheet_names": sheet_names,
                "total_rows": total_rows,
                "max_columns": total_cols,
                "title": wb.properties.title if wb.properties else "",
                "author": wb.properties.creator if wb.properties else "",
            }
            wb.close()
            return result
        except Exception as e:
            return {"error": str(e), "sheet_count": 0, "total_rows": 0}


excel_extractor = ExcelExtractor()
